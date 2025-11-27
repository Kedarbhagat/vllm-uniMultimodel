import openpyxl
import pandas as pd
import os
import logging
from typing import List, Dict, Any, Optional
from langchain_core.documents import Document
import time

# Try to import xlrd for direct .xls support
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

logger = logging.getLogger("excel_processor")

# Constants
MAX_EXCEL_SIZE_MB = 100
MAX_ROWS_PER_SHEET = 100000

class ExcelProcessingError(Exception):
    """Custom exception for Excel processing errors"""
    pass

def is_excel_safe(excel_path: str) -> bool:
    """Check if Excel file is safe to process"""
    try:
        file_size_mb = os.path.getsize(excel_path) / (1024 * 1024)
        if file_size_mb > MAX_EXCEL_SIZE_MB:
            logger.warning(f"Excel file too large: {file_size_mb}MB > {MAX_EXCEL_SIZE_MB}MB")
            return False
        return True
    except Exception as e:
        logger.error(f"Error checking Excel safety: {e}")
        return False

def extract_text_from_excel(excel_path: str, 
                            include_formulas: bool = False,
                            max_rows_per_chunk: int = 50) -> List[Dict[str, Any]]:
    """Extract text from Excel file (.xlsx, .xls) with chunking"""
    start_time = time.time()
    extracted_chunks = []
    
    try:
        file_ext = os.path.splitext(excel_path)[1].lower()
        logger.info(f"Processing Excel file with extension: {file_ext}")
        
        if file_ext == '.xlsx':
            # Use openpyxl for .xlsx files (better for formatting)
            logger.info("Using openpyxl for .xlsx file")
            extracted_chunks = _extract_with_openpyxl(
                excel_path, 
                include_formulas, 
                max_rows_per_chunk
            )
        elif file_ext == '.xls':
            # Try direct xlrd first (avoids pandas version conflicts)
            if XLRD_AVAILABLE:
                logger.info("Using xlrd directly for .xls file")
                extracted_chunks = _extract_with_xlrd(
                    excel_path,
                    max_rows_per_chunk
                )
            else:
                # Fallback to pandas
                logger.info("xlrd not available, using pandas for .xls file")
                extracted_chunks = _extract_with_pandas(
                    excel_path, 
                    max_rows_per_chunk
                )
        else:
            logger.error(f"Unsupported Excel format: {file_ext}")
            return []
        
        processing_time = time.time() - start_time
        logger.info(f"Processed Excel in {processing_time:.2f}s: {excel_path}, extracted {len(extracted_chunks)} chunks")
        return extracted_chunks
        
    except Exception as e:
        logger.exception(f"Failed to process Excel {excel_path}")
        return []

def _extract_with_openpyxl(excel_path: str, 
                           include_formulas: bool = False,
                           max_rows_per_chunk: int = 50) -> List[Dict[str, Any]]:
    """Extract using openpyxl (for .xlsx files)"""
    extracted_chunks = []
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=not include_formulas)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Skip empty sheets
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue
            
            # Check row count
            if sheet.max_row > MAX_ROWS_PER_SHEET:
                logger.warning(f"Sheet '{sheet_name}' has too many rows ({sheet.max_row}), skipping")
                continue
            
            # Extract header row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            
            # Process data in chunks
            current_chunk = []
            chunk_data = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Convert row to list of strings
                row_data = [str(cell) if cell is not None else "" for cell in row]
                
                # Skip completely empty rows
                if not any(cell.strip() for cell in row_data):
                    continue
                
                # Create row text
                row_text = " | ".join([
                    f"{headers[i]}: {row_data[i]}" 
                    for i in range(min(len(headers), len(row_data)))
                    if row_data[i].strip()
                ])
                
                chunk_data.append(row_text)
                
                # Create chunk when reaching max size
                if len(chunk_data) >= max_rows_per_chunk:
                    extracted_chunks.append({
                        "text": "\n".join(chunk_data),
                        "sheet": sheet_name,
                        "row_range": f"{row_idx - len(chunk_data) + 1}-{row_idx}",
                        "source": "openpyxl",
                        "headers": headers
                    })
                    chunk_data = []
            
            # Add remaining data as final chunk
            if chunk_data:
                extracted_chunks.append({
                    "text": "\n".join(chunk_data),
                    "sheet": sheet_name,
                    "row_range": f"{sheet.max_row - len(chunk_data) + 1}-{sheet.max_row}",
                    "source": "openpyxl",
                    "headers": headers
                })
        
        return extracted_chunks
        
    except Exception as e:
        logger.error(f"Error in openpyxl extraction: {str(e)}")
        return []

def _extract_with_xlrd(excel_path: str,
                       max_rows_per_chunk: int = 50) -> List[Dict[str, Any]]:
    """Extract using xlrd directly (for .xls files, bypasses pandas version conflicts)"""
    extracted_chunks = []
    
    try:
        logger.info(f"Opening .xls file with xlrd: {excel_path}")
        workbook = xlrd.open_workbook(excel_path)
        
        logger.info(f"Found {workbook.nsheets} sheets")
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            
            logger.info(f"Processing sheet: '{sheet_name}' with {sheet.nrows} rows and {sheet.ncols} columns")
            
            # Skip empty sheets
            if sheet.nrows == 0 or sheet.ncols == 0:
                logger.warning(f"Sheet '{sheet_name}' is empty, skipping")
                continue
            
            # Check row count
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                logger.warning(f"Sheet '{sheet_name}' has too many rows ({sheet.nrows}), skipping")
                continue
            
            # Extract headers from first row
            headers = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col_idx)
                headers.append(str(cell_value) if cell_value else f"Column{col_idx+1}")
            
            logger.info(f"Headers: {headers[:5]}...")
            
            # Process data rows in chunks
            chunk_data = []
            
            for row_idx in range(1, sheet.nrows):  # Start from 1 to skip header
                row_values = []
                
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    
                    # Convert cell value to string
                    if cell_value == "":
                        continue
                    
                    # Handle different cell types
                    if sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                        # Convert date to string
                        cell_str = str(cell_value)
                    else:
                        cell_str = str(cell_value).strip()
                    
                    if cell_str:
                        row_values.append(f"{headers[col_idx]}: {cell_str}")
                
                if row_values:
                    row_text = " | ".join(row_values)
                    chunk_data.append(row_text)
                
                # Create chunk when reaching max size
                if len(chunk_data) >= max_rows_per_chunk:
                    logger.info(f"Created chunk from rows {row_idx - len(chunk_data) + 2}-{row_idx + 1} with {len(chunk_data)} rows")
                    extracted_chunks.append({
                        "text": "\n".join(chunk_data),
                        "sheet": sheet_name,
                        "row_range": f"{row_idx - len(chunk_data) + 2}-{row_idx + 1}",
                        "source": "xlrd",
                        "headers": headers
                    })
                    chunk_data = []
            
            # Add remaining data as final chunk
            if chunk_data:
                logger.info(f"Created final chunk with {len(chunk_data)} rows")
                extracted_chunks.append({
                    "text": "\n".join(chunk_data),
                    "sheet": sheet_name,
                    "row_range": f"{sheet.nrows - len(chunk_data) + 1}-{sheet.nrows}",
                    "source": "xlrd",
                    "headers": headers
                })
        
        logger.info(f"Total chunks extracted with xlrd: {len(extracted_chunks)}")
        return extracted_chunks
        
    except Exception as e:
        logger.exception(f"Error in xlrd extraction")
        return []

def _extract_with_pandas(excel_path: str, 
                        max_rows_per_chunk: int = 50) -> List[Dict[str, Any]]:
    """Extract using pandas (works for .xls and .xlsx)"""
    extracted_chunks = []
    
    try:
        logger.info(f"Attempting to read Excel file with pandas: {excel_path}")
        
        file_ext = os.path.splitext(excel_path)[1].lower()
        
        # Try different engines for .xls files
        if file_ext == '.xls':
            logger.info("Detected .xls file, trying multiple engines...")
            
            # Try xlrd first (old Excel format)
            try:
                excel_file = pd.ExcelFile(excel_path, engine='xlrd')
                logger.info("Successfully opened with xlrd engine")
            except Exception as e1:
                logger.warning(f"xlrd failed: {str(e1)}")
                
                # Try openpyxl as fallback (sometimes works with .xls)
                try:
                    excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
                    logger.info("Successfully opened with openpyxl engine")
                except Exception as e2:
                    logger.warning(f"openpyxl failed: {str(e2)}")
                    
                    # Last resort: try without specifying engine
                    try:
                        excel_file = pd.ExcelFile(excel_path)
                        logger.info("Successfully opened with default engine")
                    except Exception as e3:
                        logger.error(f"All engines failed. xlrd: {e1}, openpyxl: {e2}, default: {e3}")
                        logger.error("For .xls files, you may need to: pip uninstall xlrd && pip install xlrd==1.2.0")
                        return []
        else:
            # For .xlsx files, use default (openpyxl)
            excel_file = pd.ExcelFile(excel_path)
            logger.info("Successfully opened .xlsx file")
        
        logger.info(f"Found {len(excel_file.sheet_names)} sheets: {excel_file.sheet_names}")
        
        for sheet_name in excel_file.sheet_names:
            try:
                logger.info(f"Processing sheet: '{sheet_name}'")
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                logger.info(f"Sheet '{sheet_name}' has {len(df)} rows and {len(df.columns)} columns")
                
                # Skip empty sheets
                if df.empty:
                    logger.warning(f"Sheet '{sheet_name}' is empty, skipping")
                    continue
                
                # Check row count
                if len(df) > MAX_ROWS_PER_SHEET:
                    logger.warning(f"Sheet '{sheet_name}' has too many rows ({len(df)}), skipping")
                    continue
                
                # Get headers
                headers = df.columns.tolist()
                logger.info(f"Headers: {headers[:5]}...")  # Show first 5 headers
                
                # Process in chunks
                for chunk_start in range(0, len(df), max_rows_per_chunk):
                    chunk_end = min(chunk_start + max_rows_per_chunk, len(df))
                    chunk_df = df.iloc[chunk_start:chunk_end]
                    
                    # Convert chunk to text
                    chunk_text = []
                    for idx, row in chunk_df.iterrows():
                        row_text = " | ".join([
                            f"{col}: {row[col]}" 
                            for col in headers 
                            if pd.notna(row[col]) and str(row[col]).strip()
                        ])
                        if row_text:
                            chunk_text.append(row_text)
                    
                    if chunk_text:
                        logger.info(f"Created chunk from rows {chunk_start + 2}-{chunk_end + 1} with {len(chunk_text)} rows")
                        extracted_chunks.append({
                            "text": "\n".join(chunk_text),
                            "sheet": sheet_name,
                            "row_range": f"{chunk_start + 2}-{chunk_end + 1}",
                            "source": "pandas",
                            "headers": headers
                        })
                
            except Exception as e:
                logger.exception(f"Error processing sheet '{sheet_name}'")
                continue
        
        logger.info(f"Total chunks extracted: {len(extracted_chunks)}")
        return extracted_chunks
        
    except Exception as e:
        logger.exception(f"Error in pandas extraction")
        return []

def convert_excel_to_langchain_docs(
    excel_path: str,
    max_rows_per_chunk: int = 50,
    include_formulas: bool = False,
    max_sentences: int = 10  # Added for compatibility with process_document call
) -> List[Document]:
    """Process Excel files to LangChain Documents
    
    Note: max_sentences parameter is included for API compatibility with other
    document processors but is not used for Excel files (we chunk by rows instead).
    """
    start_time = time.time()
    logger.info(f"Starting processing of Excel file: {excel_path}")
    
    if not is_excel_safe(excel_path):
        raise ExcelProcessingError(f"Excel file failed safety checks: {excel_path}")
    
    try:
        # Extract data from Excel
        extracted = extract_text_from_excel(
            excel_path, 
            include_formulas=include_formulas,
            max_rows_per_chunk=max_rows_per_chunk
        )
        
        documents = []
        filename = os.path.basename(excel_path)
        
        # Create LangChain documents
        for chunk_idx, entry in enumerate(extracted):
            if not entry["text"].strip():
                continue
            
            # Create unique ID for the chunk
            chunk_id = f"{filename}_{entry['sheet']}_r{entry['row_range']}_c{chunk_idx}"
            
            # Convert headers list to comma-separated string for ChromaDB compatibility
            headers_str = ", ".join(entry.get("headers", [])) if entry.get("headers") else ""
            
            # Create LangChain Document
            doc = Document(
                page_content=entry["text"],
                metadata={
                    "source": excel_path,
                    "filename": filename,
                    "sheet": entry["sheet"],
                    "row_range": entry["row_range"],
                    "chunk_id": chunk_id,
                    "extraction_method": entry["source"],
                    "headers": headers_str,  # Convert list to string
                    "processing_time": time.time() - start_time
                }
            )
            documents.append(doc)
        
        logger.info(f"Completed processing Excel file {excel_path}: {len(documents)} chunks created")
        return documents
        
    except Exception as e:
        logger.error(f"Failed to process Excel file {excel_path}: {str(e)}")
        raise ExcelProcessingError(f"Excel conversion failed: {str(e)}")